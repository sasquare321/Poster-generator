from __future__ import annotations

import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional, Tuple

from pymongo import MongoClient
from openpyxl import load_workbook


# =========================
# CONFIG
# =========================

MONGO_URI = "mongodb+srv://YOUR_USERNAME:YOUR_PASSWORD@YOUR_CLUSTER.mongodb.net/?appName=YourApp"
DB_NAME = "posters"
COLLECTION_NAME = "prices"

EXCEL_FILE = r"C:\path\to\your\prices.xlsx"
SHEET_NAME = "Sheet1"

# Expected Excel headers in row 1:
# | date | gold | silver |
#
# Example:
# | 2026-03-09 | 17000 | 315 |
#
# If date is empty, today's date will be used.

DATE_HEADER = "date"
GOLD_HEADER = "gold"
SILVER_HEADER = "silver"

# True  -> one document per date
# False -> always overwrite the single latest document
UPSERT_BY_DATE = False


# =========================
# HELPERS
# =========================

def log(message: str) -> None:
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{now}] {message}")


def normalize_header(value: object) -> str:
    return str(value).strip().lower() if value is not None else ""


def parse_date(date_value: object) -> str:
    if date_value is None or str(date_value).strip() == "":
        return datetime.now().strftime("%Y-%m-%d")

    if isinstance(date_value, datetime):
        return date_value.strftime("%Y-%m-%d")

    text = str(date_value).strip()

    # Try common formats
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass

    # If already some other readable date string, keep as-is
    return text


def parse_number(value: object, field_name: str) -> float:
    if value is None or str(value).strip() == "":
        raise ValueError(f"{field_name} value is empty")

    text = str(value).strip().replace(",", "").replace("₹", "").replace("/-", "")
    try:
        num = float(text)
    except ValueError as exc:
        raise ValueError(f"Invalid {field_name} value: {value}") from exc

    # Convert whole-number floats like 17000.0 -> 17000
    if num.is_integer():
        return int(num)
    return num


def get_header_map(sheet) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for col_idx, cell in enumerate(sheet[1], start=1):
        header = normalize_header(cell.value)
        if header:
            header_map[header] = col_idx
    return header_map


def find_last_filled_row(sheet, gold_col: int, silver_col: int) -> Optional[int]:
    for row_idx in range(sheet.max_row, 1, -1):
        gold_val = sheet.cell(row=row_idx, column=gold_col).value
        silver_val = sheet.cell(row=row_idx, column=silver_col).value
        if gold_val not in (None, "") or silver_val not in (None, ""):
            return row_idx
    return None


def read_latest_price_row(
    excel_file: str,
    sheet_name: str,
) -> Tuple[str, float, float]:
    path = Path(excel_file)

    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_file}")

    wb = load_workbook(filename=excel_file, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    sheet = wb[sheet_name]
    header_map = get_header_map(sheet)

    required = [DATE_HEADER, GOLD_HEADER, SILVER_HEADER]
    missing = [h for h in required if h not in header_map and h != DATE_HEADER]

    if missing:
        raise ValueError(
            f"Missing required headers: {missing}. "
            f"Expected headers in row 1: date, gold, silver"
        )

    date_col = header_map.get(DATE_HEADER)
    gold_col = header_map[GOLD_HEADER]
    silver_col = header_map[SILVER_HEADER]

    last_row = find_last_filled_row(sheet, gold_col, silver_col)
    if last_row is None:
        raise ValueError("No price rows found in Excel")

    raw_date = sheet.cell(row=last_row, column=date_col).value if date_col else None
    raw_gold = sheet.cell(row=last_row, column=gold_col).value
    raw_silver = sheet.cell(row=last_row, column=silver_col).value

    date_str = parse_date(raw_date)
    gold = parse_number(raw_gold, "gold")
    silver = parse_number(raw_silver, "silver")

    return date_str, gold, silver


def update_mongodb(date_str: str, gold: float, silver: float) -> None:
    client = MongoClient(MONGO_URI)
    db = client[DB_NAME]
    collection = db[COLLECTION_NAME]

    payload = {
        "gold": gold,
        "silver": silver,
        "date": date_str,
        "updatedAt": datetime.now(),
    }

    if UPSERT_BY_DATE:
        result = collection.update_one(
            {"date": date_str},
            {"$set": payload},
            upsert=True,
        )
    else:
        result = collection.update_one(
            {},
            {"$set": payload},
            upsert=True,
        )

    log(
        f"MongoDB updated successfully | "
        f"matched={result.matched_count}, modified={result.modified_count}, "
        f"upserted_id={result.upserted_id}"
    )
    client.close()


def main() -> int:
    try:
        log("Script started")

        date_str, gold, silver = read_latest_price_row(EXCEL_FILE, SHEET_NAME)
        log(f"Read from Excel -> date={date_str}, gold={gold}, silver={silver}")

        update_mongodb(date_str, gold, silver)

        log("Done")
        return 0

    except Exception as exc:
        log(f"ERROR: {exc}")
        return 1


if __name__ == "__main__":
    sys.exit(main())